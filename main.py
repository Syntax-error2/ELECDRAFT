"""
ELECDRAFT PRO – main application entry-point
=============================================
Professional Electrical CAD with PEC-based automatic load scheduling.

Supported floorplan formats (import pipeline):
    • PNG / JPG / JPEG / BMP / TIFF  – direct pixel load via Pillow
    • DXF                            – vector parse via ezdxf → Pillow rasterise
    • DWG                            – LibreCAD headless convert → DXF → rasterise
                                       (falls back to a user-friendly error if
                                        LibreCAD is not installed)

Dependencies that are imported lazily (only when the user actually hits
"Import") so that the app starts instantly even if they are missing:
    ezdxf   – pip install ezdxf
    PIL     – pip install Pillow        (usually already present)
"""

from __future__ import annotations

import sys
import os
import json
import subprocess
import tempfile
import traceback
import shutil
from typing import Any

import openpyxl

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QPushButton, QWidget,
    QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QLabel,
    QLineEdit, QGroupBox, QMessageBox, QFrame, QTabWidget, QScrollArea,
    QTreeWidget, QTreeWidgetItem, QSplitter, QMenu, QFileDialog,
    QGridLayout, QToolButton, QSizePolicy, QInputDialog, QProgressBar,
    QDialog, QDialogButtonBox,
)
from PySide6.QtGui import (
    QColor, QFont, QIcon, QAction, QPixmap, QPainter, QPen,
)
from PySide6.QtCore import (
    Qt, QParallelAnimationGroup, QPropertyAnimation, QAbstractAnimation,
    QTimer, QPointF, QThread, Signal, QRect,
)
from PySide6.QtPrintSupport import QPrinter

# ---------------------------------------------------------------------------
# Internal module imports  (your project structure)
# ---------------------------------------------------------------------------
from ui.canvas import DesignCanvas
from ui.sld_viewer import SLDViewer
from ui.view_3d import View3D
from ui.settings_dialog import ProjectSettingsDialog
from modules.logic import PECCalculator


# ==========================================================================
# CONSTANTS  –  centralised so tweaks don't require a grep
# ==========================================================================
LOGO_PATH: str = r"assets/symbols/ELECDRAFT_LOGO.png"
DEFAULT_COMPONENTS_PATH: str = "data/components.json"
SYMBOL_BASE: str = "assets/symbols/"

# File-dialog filter strings
_FLOORPLAN_FILTER = (
    "Floorplan Files (*.png *.jpg *.jpeg *.bmp *.tiff *.dxf *.dwg);;"
    "Images (*.png *.jpg *.jpeg *.bmp *.tiff);;"
    "DXF Files (*.dxf);;"
    "DWG Files (*.dwg)"
)
_PROJECT_FILTER  = "ElecDraft Project (*.json)"
_EXCEL_FILTER    = "Excel Workbook (*.xlsx)"
_PDF_FILTER      = "PDF Files (*.pdf)"
_SVG_FILTER      = "SVG Files (*.svg)"
_SYMBOL_FILTER   = "Symbol Files (*.svg *.png *.jpg *.jpeg *.bmp)"

# Colour tokens
CLR_ACCENT        = "#00e5ff"
CLR_ACCENT_GREEN  = "#00ff88"
CLR_BG_MAIN       = "#0d0f14"
CLR_BG_PANEL      = "#15191e"
CLR_BG_CARD       = "#1c222d"
CLR_BORDER        = "#2d3646"
CLR_BORDER_LIGHT  = "#232931"
CLR_TEXT_DIM      = "#505f73"
CLR_TEXT_BODY     = "#a0a0a0"
CLR_TEXT_BRIGHT   = "#e0e0e0"


# ==========================================================================
# FLOORPLAN IMPORT  –  unified DXF / DWG / PNG pipeline
# ==========================================================================
class _ImportWorker(QThread):
    """Background worker that performs the potentially-slow file conversion.

    Signals
    -------
    progress(int)       – 0-100 progress ticks
    finished_ok(str)    – path to the final PNG on success
    finished_err(str)   – human-readable error message on failure
    """

    progress     = Signal(int)
    finished_ok  = Signal(str)
    finished_err = Signal(str)

    def __init__(self, source_path: str) -> None:
        super().__init__()
        self.source_path = source_path

    # ── main dispatch ─────────────────────────────────────────────────
    def run(self) -> None:
        ext = os.path.splitext(self.source_path)[1].lower()
        try:
            if ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff"):
                self._handle_image()
            elif ext == ".dxf":
                self._handle_dxf()
            elif ext == ".dwg":
                self._handle_dwg()
            else:
                self.finished_err.emit(f"Unsupported file extension: {ext}")
        except Exception:
            self.finished_err.emit(f"Unexpected error:\n{traceback.format_exc()}")

    # ── IMAGE  (PNG / JPG / …) ────────────────────────────────────────
    def _handle_image(self) -> None:
        try:
            from PIL import Image
        except ImportError:
            self.finished_err.emit(
                "Pillow is not installed.\n\nRun:  pip install Pillow"
            )
            return

        self.progress.emit(20)
        img = Image.open(self.source_path).convert("RGBA")
        self.progress.emit(60)

        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        img.save(tmp.name, "PNG")
        tmp.close()
        self.progress.emit(100)
        self.finished_ok.emit(tmp.name)

    # ── DXF  –  vector → raster via ezdxf + Pillow ───────────────────
    def _handle_dxf(self) -> None:
        try:
            import ezdxf
        except ImportError:
            self.finished_err.emit(
                "ezdxf is not installed – it is required to read DXF files.\n\n"
                "Run:  pip install ezdxf\n\n"
                "Alternatively you can import the floorplan as a PNG or JPG image."
            )
            return

        try:
            from PIL import Image, ImageDraw
        except ImportError:
            self.finished_err.emit("Pillow is not installed.\n\nRun:  pip install Pillow")
            return

        self.progress.emit(10)
        doc = ezdxf.readfile(self.source_path)
        msp = doc.modelspace()
        self.progress.emit(30)

        # ── lightweight rasterisation ─────────────────────────────────
        # Handles LINE, LWPOLYLINE, POLYLINE, CIRCLE – the vast majority
        # of typical floorplans.  No matplotlib / ezdxf Renderer needed.
        W, H = 2400, 1800
        img  = Image.new("RGBA", (W, H), (20, 25, 35, 255))
        draw = ImageDraw.Draw(img)

        # 1) Bounding-box pass
        xs, ys = [], []
        for e in msp:
            t = e.dxftype()
            if t == "LINE":
                xs.extend([e.dxf.start.x, e.dxf.end.x])
                ys.extend([e.dxf.start.y, e.dxf.end.y])
            elif t == "LWPOLYLINE":
                for pt in e.get_points(format="xy"):
                    xs.append(pt[0]); ys.append(pt[1])
            elif t == "CIRCLE":
                cx, cy, r = e.dxf.center.x, e.dxf.center.y, e.dxf.radius
                xs.extend([cx - r, cx + r]); ys.extend([cy - r, cy + r])
            elif t == "POLYLINE":
                try:
                    for v in e.vertices:
                        xs.append(v.dxf.location.x); ys.append(v.dxf.location.y)
                except Exception:
                    pass

        self.progress.emit(45)

        if not xs or not ys:
            # Totally empty drawing – return blank dark canvas
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(tmp.name, "PNG"); tmp.close()
            self.progress.emit(100)
            self.finished_ok.emit(tmp.name)
            return

        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        rx = max_x - min_x or 1.0
        ry = max_y - min_y or 1.0
        PAD   = 60
        scale = min((W - 2*PAD) / rx, (H - 2*PAD) / ry)

        def w2p(wx, wy):          # world → pixel (Y-flip)
            return (PAD + (wx - min_x)*scale,
                    H   - PAD - (wy - min_y)*scale)

        # 2) Draw pass
        WALL = (180, 200, 220, 255)
        THIN = ( 80, 120, 160, 255)
        LW   = 3

        for e in msp:
            t = e.dxftype()
            if t == "LINE":
                draw.line([w2p(e.dxf.start.x, e.dxf.start.y),
                           w2p(e.dxf.end.x,   e.dxf.end.y)],   fill=WALL, width=LW)
            elif t == "LWPOLYLINE":
                pts = [w2p(p[0], p[1]) for p in e.get_points(format="xy")]
                if len(pts) >= 2:
                    draw.line(pts, fill=WALL, width=LW)
                    if e.closed and len(pts) >= 3:
                        draw.line([pts[-1], pts[0]], fill=WALL, width=LW)
            elif t == "CIRCLE":
                cx, cy = w2p(e.dxf.center.x, e.dxf.center.y)
                r = e.dxf.radius * scale
                draw.ellipse([cx-r, cy-r, cx+r, cy+r], outline=THIN, width=2)
            elif t == "POLYLINE":
                try:
                    pts = [w2p(v.dxf.location.x, v.dxf.location.y) for v in e.vertices]
                    if len(pts) >= 2:
                        draw.line(pts, fill=WALL, width=LW)
                except Exception:
                    pass

        self.progress.emit(85)
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        img.save(tmp.name, "PNG"); tmp.close()
        self.progress.emit(100)
        self.finished_ok.emit(tmp.name)

    # ── DWG  –  LibreCAD/LibreOffice headless → DXF → rasterise ──────
    def _handle_dwg(self) -> None:
        self.progress.emit(5)

        # Locate a headless converter (LibreOffice or LibreCAD)
        candidates = [
            "soffice",
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            "/usr/bin/libreoffice",
            "/usr/local/bin/libreoffice",
        ]
        soffice = None
        for c in candidates:
            try:
                subprocess.run([c, "--version"], capture_output=True, timeout=5)
                soffice = c
                break
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue

        if soffice is None:
            self.finished_err.emit(
                "DWG conversion requires LibreCAD or LibreOffice to be installed.\n\n"
                "1) Install from  https://www.libreCad.org\n"
                "   or            https://www.libreoffice.org\n\n"
                "2) Make sure it is on your system PATH.\n\n"
                "Alternatively, convert the DWG to DXF/PNG externally and import that."
            )
            return

        self.progress.emit(15)
        out_dir = tempfile.mkdtemp()
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "dxf",
                 "--outdir", out_dir, self.source_path],
                capture_output=True, timeout=60,
            )
            self.progress.emit(40)

            if result.returncode != 0:
                self.finished_err.emit(
                    f"Conversion failed (exit {result.returncode}).\n"
                    f"stderr: {result.stderr.decode(errors='replace')}"
                )
                return

            base   = os.path.splitext(os.path.basename(self.source_path))[0]
            dxf_path = os.path.join(out_dir, base + ".dxf")
            if not os.path.exists(dxf_path):
                self.finished_err.emit("Converter did not produce a DXF output file.")
                return

            self.progress.emit(45)
            self.source_path = dxf_path   # re-point and run DXF pipeline
            self._handle_dxf()

        except subprocess.TimeoutExpired:
            self.finished_err.emit("Conversion timed out after 60 s.")


# ── thin progress dialog ──────────────────────────────────────────────────
class _ImportProgressDialog(QDialog):
    """Dark-themed modal with a progress bar driven by the worker thread."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Importing Floorplan…")
        self.setMinimumWidth(400)
        self.setStyleSheet(f"""
            QDialog         {{ background: {CLR_BG_MAIN}; color: {CLR_TEXT_BODY}; border-radius: 8px; }}
            QLabel          {{ color: {CLR_TEXT_BRIGHT}; font-size: 12px; }}
            QProgressBar    {{ border: 1px solid {CLR_BORDER}; border-radius: 4px;
                               background: {CLR_BG_PANEL}; height: 22px; }}
            QProgressBar::chunk {{ background: {CLR_ACCENT}; border-radius: 3px; }}
            QPushButton     {{ background: {CLR_BG_CARD}; color: {CLR_TEXT_BRIGHT};
                               border: 1px solid {CLR_BORDER}; border-radius: 4px; padding: 5px 16px; }}
            QPushButton:hover {{ border-color: {CLR_ACCENT}; }}
        """)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(12)

        self.lbl = QLabel("Preparing…")
        lay.addWidget(self.lbl)

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setTextVisible(False)
        lay.addWidget(self.bar)

        cancel = QPushButton("Cancel")
        cancel.setFixedHeight(30)
        cancel.clicked.connect(self.reject)
        lay.addWidget(cancel, alignment=Qt.AlignRight)

        # result holders
        self.result_png: str | None  = None
        self.error_msg:  str | None  = None

    # ── slots ─────────────────────────────────────────────────────────
    def on_progress(self, v):
        self.bar.setValue(v)
        if v >= 100:
            self.lbl.setText("Done.")

    def on_ok(self, path):
        self.result_png = path
        self.accept()

    def on_err(self, msg):
        self.error_msg = msg
        self.reject()


# ── façade used by the main window ────────────────────────────────────────
class FloorplanImporter:
    """Kicks off the threaded import and blocks on a modal dialog."""

    def __init__(self, canvas: DesignCanvas, parent: QWidget) -> None:
        self._canvas  = canvas
        self._parent  = parent

    def run(self, path: str) -> bool:
        self._worker = _ImportWorker(path)
        self._dialog = _ImportProgressDialog(self._parent)

        self._worker.progress.connect(self._dialog.on_progress)
        self._worker.finished_ok.connect(self._dialog.on_ok)
        self._worker.finished_err.connect(self._dialog.on_err)

        self._worker.start()
        accepted = self._dialog.exec()
        self._worker.wait()          # ensure thread is dead before reading results

        if accepted and self._dialog.result_png:
            self._canvas.load_from_png(self._dialog.result_png)
            try:
                os.unlink(self._dialog.result_png)
            except OSError:
                pass
            return True

        if self._dialog.error_msg:
            QMessageBox.critical(self._parent, "Import Failed", self._dialog.error_msg)
        return False


# ==========================================================================
# COLLAPSIBLE BOX  –  animated sidebar section
# ==========================================================================
class CollapsibleBox(QWidget):
    """Expandable / collapsible panel with a smooth height animation."""

    def __init__(self, title: str = "", parent=None) -> None:
        super().__init__(parent)

        self.toggle_button = QToolButton(text=title, checkable=True, checked=True)
        self.toggle_button.setStyleSheet(f"""
            QToolButton       {{ border: none; font-weight: bold; color: {CLR_ACCENT};
                                background-color: {CLR_BG_CARD}; text-align: left;
                                padding: 8px; font-size: 11px; border-radius: 4px; }}
            QToolButton:hover {{ background-color: #252e3e; }}
        """)
        self.toggle_button.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.toggle_button.setArrowType(Qt.DownArrow)
        self.toggle_button.pressed.connect(self._on_toggle)
        self.toggle_button.setFixedHeight(34)
        self.toggle_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self._content = QWidget()
        self._content.setMaximumHeight(0)
        self._content.setMinimumHeight(0)

        self._anim = QParallelAnimationGroup(self)
        self._anim.addAnimation(QPropertyAnimation(self,          b"minimumHeight"))
        self._anim.addAnimation(QPropertyAnimation(self,          b"maximumHeight"))
        self._anim.addAnimation(QPropertyAnimation(self._content, b"maximumHeight"))

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        lay.addWidget(self.toggle_button)
        lay.addWidget(self._content)

    def _on_toggle(self):
        checked = self.toggle_button.isChecked()
        self.toggle_button.setArrowType(Qt.DownArrow if not checked else Qt.RightArrow)
        self._anim.setDirection(
            QAbstractAnimation.Forward if not checked else QAbstractAnimation.Backward
        )
        self._anim.start()

    def set_content_layout(self, layout):
        self._content.setLayout(layout)
        collapsed_h = self.sizeHint().height() - self._content.maximumHeight()
        content_h   = layout.sizeHint().height()

        for i in range(self._anim.animationCount()):
            a = self._anim.animationAt(i)
            a.setDuration(280)
            a.setStartValue(collapsed_h if i < 2 else 0)
            a.setEndValue((collapsed_h + content_h) if i < 2 else content_h)

        if self.toggle_button.isChecked():
            self._content.setMaximumHeight(content_h)
            self.setMinimumHeight(collapsed_h + content_h)


# ==========================================================================
# MAIN APPLICATION WINDOW
# ==========================================================================
class ElecDraftApp(QMainWindow):
    """Top-level window for ELECDRAFT PRO."""

    # ── construction ──────────────────────────────────────────────────
    def __init__(self, splash=None) -> None:
        super().__init__()
        self._splash        = splash
        self._current_file  = None
        self._undo_stack    = []            # list of scene-snapshots
        self._redo_stack    = []            # for redo functionality
        self._clipboard_data = None         # internal clipboard
        self._floorplan_tabs = []          # track multiple floorplans
        self._schedule_tabs = []           # track multiple schedules
        self._is_modified = False          # track unsaved changes
        self._current_selected_item = None
        self._homerun_folders = {}         # Map homerun name to tree item

        self.setWindowOpacity(0.0)
        self.setWindowIcon(QIcon(LOGO_PATH))
        self._update_window_title()
        self.resize(1680, 960)

        self.project_data: dict[str, Any] = {
            "name":            "Main Building",
            "author":          "Lead Engineer",
            "standard":        "PEC (Philippines)",
            "export_pdf":      True,
            "system_voltage":  230,
            "transformer_kva": 50,
            "transformer_z":   0.05,
        }

        self._splash_step(10, "Initializing PEC Logic Engine…")
        self._load_component_configs()

        self._splash_step(40, "Building CAD Workspace…")
        self._setup_ui()

        self._splash_step(70, "Wiring Signals…")
        self._create_main_menu()
        self._wire_signals()

        self._splash_step(90, "Applying Theme…")
        self._apply_stylesheet()

        self._splash_step(100, "Ready.")

    # ==================================================================
    # NEW METHODS FROM ENHANCEMENTS
    # ==================================================================
    def _mark_modified(self):
        """Mark project as modified"""
        if not self._is_modified:
            self._is_modified = True
            self._update_window_title()

    def _update_window_title(self):
        """Update window title with project name and modified status"""
        title = "ELECDRAFT – Professional Electrical CAD"
        if self._current_file:
            title += f" – {os.path.basename(self._current_file)}"
        if self._is_modified:
            title += " *"
        self.setWindowTitle(title)

    def _check_unsaved_changes(self):
        """Check for unsaved changes and prompt user. Returns True if OK to proceed."""
        if not self._is_modified:
            return True

        reply = QMessageBox.question(
            self, "Unsaved Changes",
            "You have unsaved changes. Do you want to save them?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel
        )

        if reply == QMessageBox.Save:
            self.save_project()
            return True
        elif reply == QMessageBox.Discard:
            return True
        else:  # Cancel
            return False

    def save_project_as(self):
        """Save project to a new file"""
        p, _ = QFileDialog.getSaveFileName(self, "Save Project As", "", _PROJECT_FILTER)
        if not p:
            return
        self._current_file = p
        self._save_to_file(p)

    def _save_to_file(self, filepath):
        """Internal method to save project data to file"""
        from ui.canvas import ElectricalComponent
        data = {"meta": self.project_data, "items": []}
        for it in self.canvas.scene.items():
            if isinstance(it, ElectricalComponent):
                data["items"].append({
                    "name": it.name,
                    "va": it.va,
                    "x": it.pos().x(),
                    "y": it.pos().y(),
                    "type": it.comp_type,
                    "symbol": it.elementId() # Using elementId to store symbol path if needed, or we can add a property
                })

        try:
            with open(filepath, "w", encoding="utf-8") as fh:
                json.dump(data, fh, indent=4)
            self._is_modified = False
            self._update_window_title()
            self.statusBar().showMessage(f"💾  Saved → {filepath}")
            QMessageBox.information(self, "Success", f"Project saved to:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Could not save:\n\n{e}")

    def redo_last(self):
        """Redo the last undone action"""
        if not self._redo_stack:
            self.statusBar().showMessage("Nothing to redo.")
            return

        from ui.canvas import ElectricalComponent
        current = [{"name": i.name, "va": i.va, "x": i.pos().x(), "y": i.pos().y(), "type": i.comp_type}
                   for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        self._undo_stack.append(current)

        snap = self._redo_stack.pop()
        self.canvas.scene.clear()
        for d in snap:
            c = self.canvas.add_component(d["name"], {"va": d["va"], "type": d.get("type", "General")})
            c.setPos(d["x"], d["y"])
        self._sync_data()
        self._mark_modified()
        self.statusBar().showMessage("↩  Redo applied.")

    def cut_selected(self):
        """Cut selected items to clipboard"""
        self.copy_selected()
        self.delete_selected()

    def copy_selected(self):
        """Copy selected items to clipboard"""
        from ui.canvas import ElectricalComponent
        items = self.canvas.scene.selectedItems()
        if not items:
            self.statusBar().showMessage("No items selected to copy.")
            return

        self._clipboard_data = []
        for item in items:
            if isinstance(item, ElectricalComponent):
                self._clipboard_data.append({
                    "name": item.name,
                    "va": item.va,
                    "x": item.pos().x(),
                    "y": item.pos().y(),
                    "type": item.comp_type
                })
        self.statusBar().showMessage(f"Copied {len(self._clipboard_data)} item(s).")

    def paste_from_clipboard(self):
        """Paste items from clipboard"""
        if not self._clipboard_data:
            self.statusBar().showMessage("Nothing to paste.")
            return

        self._push_undo()
        offset = 20
        for data in self._clipboard_data:
            c = self.canvas.add_component(data["name"], {"va": data["va"], "type": data.get("type", "General")})
            c.setPos(data["x"] + offset, data["y"] + offset)

        self._sync_data()
        self._mark_modified()
        self.statusBar().showMessage(f"Pasted {len(self._clipboard_data)} item(s).")

    def duplicate_selected(self):
        """Duplicate selected items"""
        from ui.canvas import ElectricalComponent
        items = self.canvas.scene.selectedItems()
        if not items:
            self.statusBar().showMessage("No items selected.")
            return

        self._push_undo()
        for item in items:
            if isinstance(item, ElectricalComponent):
                c = self.canvas.add_component(item.name, {"va": item.va, "type": item.comp_type})
                c.setPos(item.pos().x() + 20, item.pos().y() + 20)

        self._sync_data()
        self._mark_modified()
        self.statusBar().showMessage(f"Duplicated {len(items)} item(s).")

    def select_all(self):
        """Select all items on canvas"""
        from ui.canvas import ElectricalComponent
        count = 0
        for item in self.canvas.scene.items():
            if isinstance(item, ElectricalComponent):
                item.setSelected(True)
                count += 1
        self.statusBar().showMessage(f"Selected {count} item(s).")

    def deselect_all(self):
        """Deselect all items"""
        self.canvas.scene.clearSelection()
        self.statusBar().showMessage("Selection cleared.")

    def calculate_total_load(self):
        """Calculate and display total connected load"""
        from ui.canvas import ElectricalComponent
        items = [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        total_va = sum(item.va for item in items)
        total_amps = total_va / self.project_data["system_voltage"]

        msg = f"""
        <h3>Total Load Calculation</h3>
        <table>
            <tr><td><b>Total Components:</b></td><td>{len(items)}</td></tr>
            <tr><td><b>Total Load:</b></td><td>{total_va} VA</td></tr>
            <tr><td><b>Total Current:</b></td><td>{total_amps:.2f} A</td></tr>
        </table>
        """
        QMessageBox.information(self, "Total Load Analysis", msg)

    def _add_new_floorplan(self):
        """Add a new floor plan to the project"""
        name, ok = QInputDialog.getText(
            self, "New Floor Plan",
            "Enter floor plan name:",
            text=f"Floor Plan {len(self._floorplan_tabs) + 2}"
        )

        if ok and name.strip():
            new_item = QTreeWidgetItem(self.bldg_item, [f"📐 {name.strip()}"])
            self.tree.expandItem(self.bldg_item)
            self._floorplan_tabs.append(name.strip())
            self.statusBar().showMessage(f"Added floor plan: {name.strip()}")
            QMessageBox.information(
                self, "Floor Plan Added",
                f"Floor plan '{name.strip()}' added.\n\n"
                "Click on it to switch to it."
            )

    def _add_new_schedule(self):
        """Add a new load schedule to the project"""
        name, ok = QInputDialog.getText(
            self, "New Load Schedule",
            "Enter schedule name:",
            text=f"Load Schedule {len(self._schedule_tabs) + 2}"
        )

        if ok and name.strip():
            new_item = QTreeWidgetItem(self.bldg_item, [f"📉 {name.strip()}"])
            self.tree.expandItem(self.bldg_item)
            self._schedule_tabs.append(name.strip())
            self.statusBar().showMessage(f"Added schedule: {name.strip()}")

    def _delete_tree_item(self, item):
        """Delete a tree item"""
        reply = QMessageBox.question(
            self, "Delete Item",
            f"Delete '{item.text(0)}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            parent = item.parent()
            if parent:
                parent.removeChild(item)
                self.statusBar().showMessage(f"Deleted: {item.text(0)}")

    def _switch_to_item(self, item):
        """Switch to a specific floor plan or schedule"""
        item_text = item.text(0)

        if "📐" in item_text:
            self.tabs.setCurrentIndex(0)
            self.statusBar().showMessage(f"Switched to: {item_text}")
        elif "📉" in item_text:
            self.tabs.setCurrentIndex(1)
            self.statusBar().showMessage(f"Switched to: {item_text}")

    def closeEvent(self, event):
        """Handle window close event"""
        if self._check_unsaved_changes():
            event.accept()
        else:
            event.ignore()

    # ── splash helper ─────────────────────────────────────────────────
    def _splash_step(self, pct, msg):
        if self._splash:
            self._splash.set_progress(pct)
            self._splash.showMessage(msg, Qt.AlignBottom | Qt.AlignLeft, QColor(CLR_ACCENT))

    # ==================================================================
    # MENU BAR - FIXED VERSION (from Claude's guide)
    # ==================================================================
    def _create_main_menu(self) -> None:
        """Create fully functional menu bar"""
        mb = self.menuBar()

        # ── FILE MENU ─────────────────────────────────────────────────
        fm = mb.addMenu("&File")

        # Create actions with proper connections
        act_new = QAction("📄  New Project", self)
        act_new.setShortcut("Ctrl+N")
        act_new.triggered.connect(self.new_project)
        fm.addAction(act_new)

        act_open = QAction("📂  Open Project", self)
        act_open.setShortcut("Ctrl+O")
        act_open.triggered.connect(self.open_project)
        fm.addAction(act_open)

        act_save = QAction("💾  Save Project", self)
        act_save.setShortcut("Ctrl+S")
        act_save.triggered.connect(self.save_project)
        fm.addAction(act_save)

        act_save_as = QAction("💾  Save Project As", self)
        act_save_as.setShortcut("Ctrl+Shift+S")
        act_save_as.triggered.connect(self.save_project_as)
        fm.addAction(act_save_as)

        fm.addSeparator()

        act_import = QAction("📐  Import Floorplan", self)
        act_import.setShortcut("Ctrl+I")
        act_import.triggered.connect(self.import_floorplan)
        fm.addAction(act_import)

        act_import_symbol = QAction("📥  Import Custom Symbol", self)
        act_import_symbol.triggered.connect(self.import_custom_symbol)
        fm.addAction(act_import_symbol)

        act_excel = QAction("📤  Export to Excel", self)
        act_excel.setShortcut("Ctrl+E")
        act_excel.triggered.connect(self.export_to_excel)
        fm.addAction(act_excel)

        act_pdf = QAction("🖨️  Export PDF Plot", self)
        act_pdf.setShortcut("Ctrl+P")
        act_pdf.triggered.connect(self.export_to_pdf)
        fm.addAction(act_pdf)

        fm.addSeparator()

        act_exit = QAction("❌  Exit", self)
        act_exit.setShortcut("Ctrl+Q")
        act_exit.triggered.connect(self.close)
        fm.addAction(act_exit)

        # ── EDIT MENU ─────────────────────────────────────────────────
        em = mb.addMenu("&Edit")

        act_undo = QAction("🔙  Undo", self)
        act_undo.setShortcut("Ctrl+Z")
        act_undo.triggered.connect(self.undo_last)
        em.addAction(act_undo)

        act_redo = QAction("🔄  Redo", self)
        act_redo.setShortcut("Ctrl+Y")
        act_redo.triggered.connect(self.redo_last)
        em.addAction(act_redo)

        em.addSeparator()

        act_cut = QAction("✂️  Cut", self)
        act_cut.setShortcut("Ctrl+X")
        act_cut.triggered.connect(self.cut_selected)
        em.addAction(act_cut)

        act_copy = QAction("📋  Copy", self)
        act_copy.setShortcut("Ctrl+C")
        act_copy.triggered.connect(self.copy_selected)
        em.addAction(act_copy)

        act_paste = QAction("📄  Paste", self)
        act_paste.setShortcut("Ctrl+V")
        act_paste.triggered.connect(self.paste_from_clipboard)
        em.addAction(act_paste)

        act_duplicate = QAction("⬡   Duplicate", self)
        act_duplicate.setShortcut("Ctrl+D")
        act_duplicate.triggered.connect(self.duplicate_selected)
        em.addAction(act_duplicate)

        em.addSeparator()

        act_select_all = QAction("🔍  Select All", self)
        act_select_all.setShortcut("Ctrl+A")
        act_select_all.triggered.connect(self.select_all)
        em.addAction(act_select_all)

        act_deselect = QAction("🔍  Deselect All", self)
        act_deselect.setShortcut("Ctrl+Shift+A")
        act_deselect.triggered.connect(self.deselect_all)
        em.addAction(act_deselect)

        em.addSeparator()

        act_delete = QAction("🗑️  Delete Selected", self)
        act_delete.setShortcut("Del")
        act_delete.triggered.connect(self.delete_selected)
        em.addAction(act_delete)

        em.addSeparator()

        act_settings = QAction("⚙️  Preferences", self)
        act_settings.triggered.connect(self._open_settings)
        em.addAction(act_settings)

        # ── ANALYSIS MENU ─────────────────────────────────────────────
        am = mb.addMenu("&Analysis")

        act_room_loads = QAction("🔍  Analyze Room Loads", self)
        act_room_loads.triggered.connect(self.run_room_analysis)
        am.addAction(act_room_loads)

        act_total_load = QAction("⚡  Calculate Total Load", self)
        act_total_load.triggered.connect(self.calculate_total_load)
        am.addAction(act_total_load)

        am.addSeparator()

        act_usability = QAction("📋  Run Usability Test", self)
        act_usability.triggered.connect(self.run_usability_evaluation)
        am.addAction(act_usability)

        # ── HELP MENU ─────────────────────────────────────────────────
        hm = mb.addMenu("&Help")

        act_about = QAction("ℹ️   About ELECDRAFT", self)
        act_about.triggered.connect(self._show_about)
        hm.addAction(act_about)

    # ==================================================================
    # SIGNAL WIRING - UPDATED
    # ==================================================================
    def _wire_signals(self) -> None:
        self.canvas.signals.circuit_updated.connect(self._sync_data)
        self.canvas.signals.circuit_updated.connect(self._mark_modified)
        self.canvas.scene.selectionChanged.connect(self._on_selection_changed)
        self.canvas.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas.customContextMenuRequested.connect(self._show_canvas_ctx)

    # ==================================================================
    # FADE-IN
    # ==================================================================
    def fade_in(self) -> None:
        self.show()
        self._fade = QPropertyAnimation(self, b"windowOpacity")
        self._fade.setDuration(700)
        self._fade.setStartValue(0.0)
        self._fade.setEndValue(1.0)
        self._fade.start()

    # ==================================================================
    # COMPONENT LIBRARY
    # ==================================================================
    def _load_component_configs(self) -> None:
        try:
            with open(DEFAULT_COMPONENTS_PATH, "r", encoding="utf-8") as fh:
                self.comp_library = json.load(fh)
        except (FileNotFoundError, json.JSONDecodeError):
            self.comp_library = {
                "💡 Light":       {"va": 100,  "is_continuous": True,  "type": "Lighting",   "symbol": SYMBOL_BASE+"light.svg"},
                "🔦 Emergency":   {"va": 50,   "is_continuous": True,  "type": "Lighting",   "symbol": SYMBOL_BASE+"emergency.svg"},
                "🏮 Chandelier":  {"va": 300,  "is_continuous": True,  "type": "Lighting",   "symbol": SYMBOL_BASE+"chandelier.svg"},
                "🔌 Duplex":      {"va": 180,  "is_continuous": False, "type": "Receptacle", "symbol": SYMBOL_BASE+"duplex.svg"},
                "🛁 GFCI Outlet": {"va": 180,  "is_continuous": False, "type": "Receptacle", "symbol": SYMBOL_BASE+"gfci.svg"},
                "🌀 Industrial":  {"va": 1000, "is_continuous": False, "type": "Receptacle", "symbol": SYMBOL_BASE+"industrial.svg"},
                "⚙️ Motor":       {"va": 1500, "is_continuous": True,  "type": "Motor",      "symbol": SYMBOL_BASE+"motor.svg"},
                "🏗️ Pump":        {"va": 2200, "is_continuous": True,  "type": "Motor",      "symbol": SYMBOL_BASE+"pump.svg"},
                "❄️ AC Unit":     {"va": 3500, "is_continuous": True,  "type": "AC",         "symbol": SYMBOL_BASE+"ac.svg"},
                "📉 Panelboard":  {"va": 0,    "is_continuous": False, "type": "Panel",      "symbol": SYMBOL_BASE+"panel.svg"},
                "🔌 Feeder":      {"va": 0,    "is_continuous": False, "type": "Feeder",     "symbol": SYMBOL_BASE+"feeder.svg"},
                "🛡️ 1-Pole":      {"va": 0,    "is_continuous": False, "type": "Breaker",    "symbol": SYMBOL_BASE+"breaker.svg"},
            }

    # ==================================================================
    # UI LAYOUT
    # ==================================================================
    def _setup_ui(self) -> None:
        self.statusBar().showMessage(f"Project: {self.project_data['name']}  │  Mode: Design")

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_header())

        self.workspace_splitter = QSplitter(Qt.Horizontal)
        self.workspace_splitter.setHandleWidth(4)
        root.addWidget(self.workspace_splitter)

        self.workspace_splitter.addWidget(self._build_left_sidebar())
        self.workspace_splitter.addWidget(self._build_centre())
        self.workspace_splitter.addWidget(self._build_right_sidebar())
        self.workspace_splitter.setStretchFactor(0, 1)
        self.workspace_splitter.setStretchFactor(1, 3)
        self.workspace_splitter.setStretchFactor(2, 1)

        # live total-load label pinned to the right of the status bar
        self._lbl_total = QLabel("Total: 0 VA")
        self._lbl_total.setStyleSheet(f"color: {CLR_ACCENT}; font-weight: bold; padding-right: 14px;")
        self.statusBar().addPermanentWidget(self._lbl_total)

    # ── header ────────────────────────────────────────────────────────
    def _build_header(self) -> QFrame:
        f = QFrame()
        f.setFixedHeight(48)
        f.setStyleSheet(f"background-color: #1a1f26; border-bottom: 1px solid {CLR_BORDER_LIGHT};")
        lay = QHBoxLayout(f)
        lay.setContentsMargins(12, 0, 12, 0)

        logo = QLabel()
        logo.setPixmap(QPixmap(LOGO_PATH).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        title = QLabel("ELECDRAFT PRO")
        title.setStyleSheet(f"color: {CLR_ACCENT}; font-weight: bold; font-size: 15px; letter-spacing: 2px;")

        lay.addWidget(logo)
        lay.addWidget(title)
        lay.addSpacing(24)

        for txt, clr, slot in [
            ("📂 FLOORPLAN",      CLR_TEXT_BODY,   self.import_floorplan),
            ("📐 IMPORT DXF/DWG", CLR_ACCENT_GREEN, self.import_floorplan),
        ]:
            b = QPushButton(txt)
            b.setStyleSheet(
                f"color:{clr}; background:transparent; border:1px solid {CLR_BORDER}; "
                f"padding:4px 12px; font-size:10px; border-radius:3px;"
            )
            b.clicked.connect(slot)
            lay.addWidget(b)

        lay.addStretch()
        return f

    # ── left sidebar ──────────────────────────────────────────────────
    def _build_left_sidebar(self) -> QFrame:
        f = QFrame()
        f.setMinimumWidth(275)
        f.setObjectName("sidePanel")
        lay = QVBoxLayout(f)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(4)

        lay.addWidget(self._sec_label("PROJECT NAVIGATION"))

        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.setFixedHeight(140)
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._tree_ctx)

        root_i  = QTreeWidgetItem(self.tree, ["📁 Site Project"])
        self.bldg_item  = QTreeWidgetItem(root_i,    [f"🏢 {self.project_data['name']}"])
        QTreeWidgetItem(self.bldg_item, ["📐 Floor Plan 1"])
        QTreeWidgetItem(self.bldg_item, ["📉 Load Schedule"])
        self.tree.expandAll()
        self.tree.itemClicked.connect(self._on_tree_click)
        lay.addWidget(self.tree)

        lay.addWidget(self._sec_label("COMPONENT TOOLBOX"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")

        tw = QWidget()
        self._tool_lay = QVBoxLayout(tw)
        self._tool_lay.setContentsMargins(0, 0, 0, 0)
        self._tool_lay.setSpacing(4)

        self._populate_toolbox()

        scroll.setWidget(tw)
        lay.addWidget(scroll, stretch=1)
        return f

    # ── centre ────────────────────────────────────────────────────────
    def _build_centre(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)

        vs = QSplitter(Qt.Vertical)

        self.tabs = QTabWidget()
        self.canvas  = DesignCanvas()
        self.view_3d = View3D()
        self.tabs.addTab(self.canvas,  "📐 Floor Plan View")
        self.tabs.addTab(QWidget(),    "📉 Load Schedule")
        self.tabs.addTab(self.view_3d, "📦 3D View")
        self.tabs.currentChanged.connect(self._on_tab)
        vs.addWidget(self.tabs)

        vs.addWidget(self._build_table_panel())
        vs.setStretchFactor(0, 7)
        vs.setStretchFactor(1, 3)

        lay.addWidget(vs)
        return w

    def _build_table_panel(self) -> QFrame:
        f = QFrame()
        f.setObjectName("tablePanel")
        lay = QVBoxLayout(f)
        lay.setContentsMargins(6, 4, 6, 4)

        hdr = QHBoxLayout()
        hdr.addWidget(self._sec_label("AUTOMATED LOAD SCHEDULE"))
        hdr.addStretch()

        # GEN SLD
        b_sld = QPushButton("📉 GEN SLD")
        b_sld.setObjectName("exportButton")
        b_sld.clicked.connect(self._open_sld)
        hdr.addWidget(b_sld)

        # DRAW CIRCUIT (toggle)
        b_wire = QPushButton("⚡ DRAW CIRCUIT")
        b_wire.setObjectName("exportButton")
        b_wire.setCheckable(True)
        b_wire.clicked.connect(lambda c: self.canvas.toggle_wire_mode(c))
        hdr.addWidget(b_wire)

        # EXPORT EXCEL
        b_xl = QPushButton("📊 EXPORT EXCEL")
        b_xl.setObjectName("exportButton")
        b_xl.clicked.connect(self.export_to_excel)
        hdr.addWidget(b_xl)

        lay.addLayout(hdr)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["#", "DESCRIPTION", "PANEL", "VOLTAGE", "LOAD (VA)", "AMPS", "WIRE SIZE"]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        lay.addWidget(self.table)
        return f

    # ── right sidebar ─────────────────────────────────────────────────
    def     _build_right_sidebar(self) -> QFrame:
        f = QFrame()
        f.setMinimumWidth(290)
        f.setObjectName("sidePanel")
        lay = QVBoxLayout(f)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(4)

        lay.addWidget(self._sec_label("PROPERTIES & PARAMETERS"))

        # Identity
        id_box = QGroupBox("Identity")
        id_l = QVBoxLayout()
        self._name_edit = QLineEdit()
        self._name_edit.setPlaceholderText("Select a component…")
        self._name_edit.textChanged.connect(self._apply_props)
        id_l.addWidget(QLabel("Power Tag:"))
        id_l.addWidget(self._name_edit)
        id_box.setLayout(id_l)
        lay.addWidget(id_box)

        # Electrical
        el_box = QGroupBox("Electrical Parameters")
        el_l = QVBoxLayout()
        self._va_edit = QLineEdit()
        self._va_edit.setPlaceholderText("0")
        self._va_edit.textChanged.connect(self._apply_props)
        el_l.addWidget(QLabel("Power (VA):"))
        el_l.addWidget(self._va_edit)
        el_l.addWidget(QLabel("Voltage: 230 V / 1 Ph"))
        el_box.setLayout(el_l)
        lay.addWidget(el_box)

        # Validation
        vb = QGroupBox("Validation & Accuracy")
        vl = QVBoxLayout()
        self._lbl_acc  = QLabel("Load Accuracy  : –")
        self._lbl_isc  = QLabel("Short Circuit  : –")
        self._lbl_vd   = QLabel("V-Drop Sync    : –")
        self._lbl_usa  = QLabel("Usability Score: –")
        for lb in (self._lbl_acc, self._lbl_isc, self._lbl_vd, self._lbl_usa):
            lb.setStyleSheet(f"color:{CLR_ACCENT_GREEN}; font-family:'Consolas'; font-size:10px;")
            vl.addWidget(lb)
        vb.setLayout(vl)
        lay.addWidget(vb)

        # Sizing summary
        lay.addWidget(self._sec_label("SIZING VERIFICATION"))
        self._summary = QLabel("Select a component\nto view sizing data…")
        self._summary.setObjectName("summaryBox")
        lay.addWidget(self._summary)

        # Buttons
        b_usa = QPushButton("📋 RUN USABILITY TEST")
        b_usa.setObjectName("toolButton")
        b_usa.clicked.connect(self.run_usability_evaluation)
        lay.addWidget(b_usa)

        lay.addStretch()

        b_set = QPushButton("⚙  PROJECT SETTINGS")
        b_set.setObjectName("toolButton")
        b_set.clicked.connect(self._open_settings)
        lay.addWidget(b_set)
        return f

    # ── shared label helper ───────────────────────────────────────────
    @staticmethod
    def _sec_label(text):
        lb = QLabel(text)
        lb.setStyleSheet(
            f"font-size:9px; font-weight:bold; color:{CLR_TEXT_DIM}; "
            f"padding:4px 0 2px 0; border:none;"
        )
        return lb

    # ==================================================================
    # FLOORPLAN IMPORT
    # ==================================================================
    def import_floorplan(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Import Floorplan", "", _FLOORPLAN_FILTER)
        if not path:
            self.statusBar().showMessage("Import cancelled.")
            return

        ok = FloorplanImporter(self.canvas, self).run(path)
        if ok:
            self.statusBar().showMessage(f"✓  Floorplan loaded: {os.path.basename(path)}")
            QMessageBox.information(
                self, "Import Successful",
                f"<b>{os.path.basename(path)}</b> loaded as the canvas background.\n\n"
                "Place electrical components on the floor plan.\n"
                "Dark areas (walls) will be avoided during wire routing.",
            )
        else:
            self.statusBar().showMessage("✗  Import failed or cancelled.")

    # ==================================================================
    # CUSTOM SYMBOL IMPORT
    # ==================================================================
    def import_custom_symbol(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Import Custom Symbol", "", _SYMBOL_FILTER)
        if not path:
            return

        try:
            # Copy to assets folder
            filename = os.path.basename(path)
            dest_path = os.path.join(SYMBOL_BASE, filename)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            
            shutil.copy2(path, dest_path)
            
            # Add to library
            name = os.path.splitext(filename)[0].capitalize()
            key = f"⭐ {name}"
            self.comp_library[key] = {
                "va": 0, 
                "is_continuous": False, 
                "type": "Custom", 
                "symbol": dest_path
            }
            
            # Save to disk
            with open(DEFAULT_COMPONENTS_PATH, "w", encoding="utf-8") as fh:
                json.dump(self.comp_library, fh, indent=4)

            # Refresh toolbox
            self._populate_toolbox()
            
            self.statusBar().showMessage(f"Imported symbol: {name}")
            QMessageBox.information(self, "Success", f"Custom symbol '{name}' imported successfully.")
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import symbol:\n{e}")

    # ==================================================================
    # PROJECT I / O - UPDATED METHODS
    # ==================================================================
    def save_project(self):
        """Save project to current file or prompt for location"""
        if not self._current_file:
            self.save_project_as()
        else:
            self._save_to_file(self._current_file)

    def new_project(self):
        """Create a new project"""
        if not self._check_unsaved_changes():
            return
        self._push_undo()
        self.canvas.scene.clear()
        self._current_file = None
        self._is_modified = False
        self._update_window_title()
        self._sync_data()
        self._reset_sidebar()
        self.statusBar().showMessage("New project created.")

    def open_project(self):
        """Open an existing project"""
        if not self._check_unsaved_changes():
            return
        p, _ = QFileDialog.getOpenFileName(self, "Open Project", "", _PROJECT_FILTER)
        if not p:
            return
        try:
            with open(p, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            self._push_undo()
            self.canvas.scene.clear()
            self.project_data.update(data.get("meta", {}))
            for d in data.get("items", []):
                c = self.canvas.add_component(d["name"], {"va": d["va"], "type": d.get("type", "General"), "symbol": d.get("symbol")})
                c.setPos(d["x"], d["y"])
            self._current_file = p
            self._is_modified = False
            self._update_window_title()
            self._sync_data()
            self.statusBar().showMessage(f"📂  Loaded: {p}")
        except (json.JSONDecodeError, KeyError, OSError) as e:
            QMessageBox.critical(self, "Load Error", f"Could not open project:\n\n{e}")

    # ==================================================================
    # UNDO/REDO - UPDATED METHODS
    # ==================================================================
    def _push_undo(self):
        from ui.canvas import ElectricalComponent
        snap = [{"name": i.name, "va": i.va, "x": i.pos().x(), "y": i.pos().y(), "type": i.comp_type}
                for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        self._undo_stack.append(snap)
        self._redo_stack.clear()  # Clear redo stack on new action
        if len(self._undo_stack) > 30:
            self._undo_stack.pop(0)

    def undo_last(self):
        if not self._undo_stack:
            self.statusBar().showMessage("Nothing to undo.")
            return

        # Save current state to redo stack
        from ui.canvas import ElectricalComponent
        current = [{"name": i.name, "va": i.va, "x": i.pos().x(), "y": i.pos().y(), "type": i.comp_type}
                   for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        self._redo_stack.append(current)

        snap = self._undo_stack.pop()
        self.canvas.scene.clear()
        for d in snap:
            c = self.canvas.add_component(d["name"], {"va": d["va"], "type": d.get("type", "General")})
            c.setPos(d["x"], d["y"])
        self._sync_data()
        self._mark_modified()
        self.statusBar().showMessage("↩  Undo applied.")

    # ==================================================================
    # ROOM MANAGEMENT
    # ==================================================================
    def _on_add_room(self) -> None:
        name, ok = QInputDialog.getText(self, "New Room", "Enter room name:")
        if ok and name.strip():
            self.canvas.add_room(name.strip())

    def run_room_analysis(self) -> None:
        from ui.canvas import ElectricalComponent
        rooms = [i for i in self.canvas.scene.items() if getattr(i, "is_room_rect", False)]
        comps = [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]

        lines = ["ROOM LOAD SUMMARY (PEC COMPLIANCE)", "=" * 44]
        total = 0
        seen  = set()
        for room in rooms:
            va = 0
            for c in comps:
                if room.contains(room.mapFromScene(c.scenePos() + QPointF(20, 20))):
                    va += c.va; seen.add(id(c))
            lines.append(f"  📍 {room.name:<28s}  {va:>6} VA")
            total += va

        orphan = sum(c.va for c in comps if id(c) not in seen)
        if orphan:
            lines.append(f"\n  ⚠️  Unassigned (outside rooms)  {orphan:>6} VA")
            total += orphan

        lines.append("=" * 44)
        lines.append(f"  TOTAL CONNECTED LOAD            {total:>6} VA")
        QMessageBox.information(self, "Room Load Analysis", "\n".join(lines))

    # ==================================================================
    # PDF EXPORT
    # ==================================================================
    def export_to_pdf(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Plot to PDF", "", _PDF_FILTER)
        if not path:
            return

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(path)
        printer.setPageOrientation(QPrinter.Landscape)
        printer.setFullPage(True)

        p = QPainter(printer)
        p.setRenderHint(QPainter.Antialiasing)

        page = printer.pageRect(QPrinter.DevicePixel)
        self.canvas.scene.render(p, page, self.canvas.scene.sceneRect())

        # ── title block ───────────────────────────────────────────────
        TW, TH = 520, 160
        tx = page.width()  - TW - 30
        ty = page.height() - TH - 30

        p.setBrush(QColor("#f5f5f0"))
        p.setPen(QPen(Qt.black, 3))
        p.drawRect(QRect(tx, ty, TW, TH))

        p.setPen(QPen(Qt.black, 2))
        p.drawLine(tx, ty + 30, tx + TW, ty + 30)

        p.setFont(QFont("Arial", 11, QFont.Bold))
        p.setPen(Qt.black)
        p.drawText(tx + 12, ty + 22, "TITLE BLOCK")

        rows = [
            ("PROJECT",  self.project_data["name"].upper()),
            ("ENGINEER", self.project_data["author"]),
            ("STANDARD", self.project_data["standard"]),
            ("VOLTAGE",  f"{self.project_data['system_voltage']} V / 1 PH"),
        ]
        for i, (k, v) in enumerate(rows):
            y = ty + 52 + i * 24
            p.setFont(QFont("Arial", 8, QFont.Bold));  p.drawText(tx+12,  y, k+":")
            p.setFont(QFont("Arial", 9));              p.drawText(tx+100, y, v)

        p.setPen(QPen(Qt.black, 8))
        p.drawRect(page)
        p.end()
        self.statusBar().showMessage(f"🖨️  PDF plotted → {path}")

    # ==================================================================
    # KEYBOARD & CONTEXT MENUS
    # ==================================================================
    def keyPressEvent(self, event) -> None:
        if event.key() in (Qt.Key_Delete, Qt.Key_Backspace):
            self.delete_selected()
        elif event.key() == Qt.Key_A and event.modifiers() == Qt.ControlModifier:
            self.select_all()
        elif event.key() == Qt.Key_A and event.modifiers() == (Qt.ControlModifier | Qt.ShiftModifier):
            self.deselect_all()
        elif event.key() == Qt.Key_D and event.modifiers() == Qt.ControlModifier:
            self.duplicate_selected()
        super().keyPressEvent(event)

    def _show_canvas_ctx(self, pos) -> None:
        item = self.canvas.itemAt(pos)
        menu = QMenu()

        if item:
            from ui.canvas import ElectricalComponent
            if isinstance(item, ElectricalComponent):
                menu.addAction("📋  Copy",             lambda: self._copy(item))
                menu.addAction("📄  Paste",            self._paste)
                menu.addAction("⬡   Duplicate",       lambda: self._duplicate(item))
                menu.addSeparator()
                menu.addAction("✏️  Edit Properties",  lambda: self._edit_props_dlg(item))
                menu.addSeparator()

        menu.addAction("🗑️  Delete Selected", self.delete_selected)
        menu.exec(self.canvas.mapToGlobal(pos))

    # ── clipboard helpers ─────────────────────────────────────────────
    def _copy(self, item):
        QApplication.clipboard().setText(json.dumps({"name": item.name, "va": item.va, "type": item.comp_type}))

    def _paste(self):
        try:
            d = json.loads(QApplication.clipboard().text())
            c = self.canvas.add_component(d["name"], {"va": d["va"], "type": d.get("type", "General")})
            c.setPos(self.canvas.mapToScene(self.canvas.viewport().rect().center()))
            self._sync_data()
            self._mark_modified()
        except (json.JSONDecodeError, KeyError):
            self.statusBar().showMessage("Nothing valid to paste.")

    def _duplicate(self, item):
        c = self.canvas.add_component(item.name, {"va": item.va, "type": item.comp_type})
        c.setPos(item.pos().x() + 20, item.pos().y() + 20)
        self._sync_data()
        self._mark_modified()

    # ── inline property editor ────────────────────────────────────────
    def _edit_props_dlg(self, item) -> None:
        from ui.canvas import ElectricalComponent
        if not isinstance(item, ElectricalComponent):
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Component")
        dlg.setMinimumWidth(300)
        dlg.setStyleSheet(f"""
            QDialog    {{ background:{CLR_BG_MAIN}; color:{CLR_TEXT_BRIGHT}; border-radius:6px; }}
            QLabel     {{ color:{CLR_TEXT_BODY}; font-size:11px; }}
            QLineEdit  {{ background:{CLR_BG_PANEL}; border:1px solid {CLR_BORDER};
                          color:#fff; padding:6px; border-radius:3px; }}
            QPushButton{{ background:{CLR_BG_CARD}; color:{CLR_TEXT_BRIGHT};
                          border:1px solid {CLR_BORDER}; border-radius:4px; padding:5px 14px; }}
            QPushButton:hover {{ border-color:{CLR_ACCENT}; }}
        """)
        l = QVBoxLayout(dlg)
        l.addWidget(QLabel("Name:"))
        ne = QLineEdit(item.name);  l.addWidget(ne)
        l.addWidget(QLabel("Power (VA):"))
        ve = QLineEdit(str(item.va)); l.addWidget(ve)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        l.addWidget(bb)

        if dlg.exec():
            try:
                item.update_data(ne.text(), int(ve.text() or 0))
                self._sync_data()
                self._mark_modified()
            except ValueError:
                pass

    def delete_selected(self) -> None:
        items = self.canvas.scene.selectedItems()
        if not items:
            return
        if QMessageBox.question(self, "Delete",
                                f"Delete {len(items)} item(s)?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self._push_undo()
            for it in items:
                self.canvas.scene.removeItem(it)
            self._sync_data()
            self._mark_modified()

    # ==================================================================
    # COMPONENT TOOLBOX
    # ==================================================================
    def _populate_toolbox(self) -> None:
        """Refreshes the component toolbox based on current library."""
        # Clear existing items in layout
        while self._tool_lay.count():
            child = self._tool_lay.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Add Room button
        btn_room = QPushButton("🏠 ADD NEW ROOM AREA")
        btn_room.setObjectName("toolButton")
        btn_room.clicked.connect(self._on_add_room)
        self._tool_lay.addWidget(btn_room)

        # Categorize items
        cats = {
            "Lighting": [],
            "Receptacles": [],
            "Loads / AC": [],
            "Distribution": [],
            "Custom": []
        }
        
        # Pre-defined order for standard items
        standard_order = [
            "💡 Light", "🔦 Emergency", "🏮 Chandelier",
            "🔌 Duplex", "🛁 GFCI Outlet", "🌀 Industrial",
            "⚙️ Motor", "🏗️ Pump", "❄️ AC Unit",
            "📉 Panelboard", "🔌 Feeder", "🛡️ 1-Pole"
        ]
        
        # Helper to find category
        def get_cat(name, data):
            t = data.get("type", "General")
            if t == "Lighting": return "Lighting"
            if t == "Receptacle": return "Receptacles"
            if t in ["Motor", "AC"]: return "Loads / AC"
            if t in ["Panel", "Feeder", "Breaker"]: return "Distribution"
            if t == "Custom": return "Custom"
            return "Custom"

        # First add standard items in order
        processed = set()
        for name in standard_order:
            if name in self.comp_library:
                cat = get_cat(name, self.comp_library[name])
                cats[cat].append(name)
                processed.add(name)
        
        # Add remaining items
        for name, data in self.comp_library.items():
            if name not in processed:
                cat = get_cat(name, data)
                cats[cat].append(name)

        # Create categories
        cat_order = ["Lighting", "Receptacles", "Loads / AC", "Distribution", "Custom"]
        
        for cat in cat_order:
            if cats[cat]:
                self._add_category(cat, cats[cat])
                
        self._tool_lay.addStretch()

    def _add_category(self, title, names):
        box = CollapsibleBox(title.upper())
        grid = QGridLayout()
        grid.setSpacing(4)
        grid.setContentsMargins(4, 4, 4, 4)

        for idx, name in enumerate(names):
            btn = QPushButton()
            btn.setFixedSize(80, 86)
            btn.setObjectName("componentTile")

            parts = name.split(" ", 1)
            ico  = parts[0] if len(parts) > 1 else "⚙"
            txt  = parts[1] if len(parts) > 1 else name

            inner = QVBoxLayout(btn)
            inner.setContentsMargins(0, 4, 0, 2)

            il = QLabel(ico);  il.setAlignment(Qt.AlignCenter)
            il.setStyleSheet("font-size:22px; color:#fff;")
            inner.addWidget(il)

            nl = QLabel(txt);  nl.setAlignment(Qt.AlignCenter)
            nl.setStyleSheet(f"font-size:8px; color:{CLR_ACCENT}; font-weight:bold;")
            inner.addWidget(nl)

            cfg = self.comp_library.get(name, {"va": 180, "type": "General"})
            btn.clicked.connect(lambda _c, n=name, c=cfg: self.canvas.add_component(n, c))

            r, c = divmod(idx, 3)
            grid.addWidget(btn, r, c)

        box.set_content_layout(grid)
        self._tool_lay.addWidget(box)

    # ==================================================================
    # TREE - ENHANCED CONTEXT MENU
    # ==================================================================
    def _tree_ctx(self, pos):
        """Enhanced context menu for project tree"""
        it = self.tree.itemAt(pos)
        if not it:
            return

        m = QMenu()
        item_text = it.text(0)

        # Building-level actions
        if "🏢" in item_text:
            m.addAction("➕ Add Floor Plan", self._add_new_floorplan)
            m.addAction("➕ Add Load Schedule", self._add_new_schedule)
            m.addSeparator()
            m.addAction("✏️ Rename Building", lambda: self.tree.editItem(it, 0))

        # Floor plan actions
        elif "📐" in item_text:
            m.addAction("🔄 Switch to this Floor Plan", lambda: self._switch_to_item(it))
            m.addAction("✏️ Rename", lambda: self.tree.editItem(it, 0))
            m.addAction("🗑️ Delete Floor Plan", lambda: self._delete_tree_item(it))

        # Schedule actions
        elif "📉" in item_text:
            m.addAction("🔄 Switch to this Schedule", lambda: self._switch_to_item(it))
            m.addAction("✏️ Rename", lambda: self.tree.editItem(it, 0))
            m.addAction("🗑️ Delete Schedule", lambda: self._delete_tree_item(it))

        m.exec(self.tree.viewport().mapToGlobal(pos))

    def _on_tree_click(self, item, _col):
        t = item.text(0)
        if "Floor Plan" in t:  self.tabs.setCurrentIndex(0)
        elif "Schedule"  in t: self.tabs.setCurrentIndex(1)

    # ==================================================================
    # TAB SWITCH
    # ==================================================================
    def _on_tab(self, idx):
        if idx == 2:
            from ui.canvas import ElectricalComponent
            self.view_3d.update_3d_scene(
                [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
            )

    # ==================================================================
    # DATA SYNC
    # ==================================================================
    def _sync_data(self) -> None:
        from ui.canvas import ElectricalComponent
        items = [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        self.table.setRowCount(len(items))

        total = 0
        
        # Clear existing homerun folders in tree
        # We will rebuild them based on current items
        # Note: This is a simplified approach. For a robust system, we should manage tree items more carefully.
        # For now, let's just ensure we have the structure.
        
        for row, item in enumerate(items):
            amps, breaker, wire, vd = PECCalculator.calculate_load(item.va)
            total += item.va

            self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.table.setItem(row, 1, QTableWidgetItem(item.name))
            self.table.setItem(row, 2, QTableWidgetItem("MAIN"))
            self.table.setItem(row, 3, QTableWidgetItem("230 V"))
            self.table.setItem(row, 4, QTableWidgetItem(f"{item.va} VA"))
            self.table.setItem(row, 5, QTableWidgetItem(f"{amps} A"))
            self.table.setItem(row, 6, QTableWidgetItem(wire))

            if item is self._current_selected_item:
                self._summary.setText(
                    f"Required Breaker : {breaker} A\n"
                    f"Required Wire    : {wire}\n"
                    f"Voltage Drop     : {vd} %"
                )
                self._lbl_acc.setText(f"Load Accuracy  : {min(100,(item.va/180)*100):.1f} %")
                self._lbl_isc.setText(f"Short Circuit  : {self._calc_isc(item.va)} kA")
                self._lbl_vd.setText(f"V-Drop Sync    : {vd} %")
            
            # Check for Homerun/Feeder to update tree
            if item.comp_type == "Feeder" or "Homerun" in item.name:
                self._update_homerun_folder(item)

        self._lbl_total.setText(f"Total: {total} VA")

    def _update_homerun_folder(self, item):
        """Updates or creates a folder in the sidebar tree for a homerun."""
        name = item.name
        if name not in self._homerun_folders:
            # Create new folder under the building item
            folder = QTreeWidgetItem(self.bldg_item, [f"⚡ {name}"])
            self._homerun_folders[name] = folder
            self.tree.expandItem(self.bldg_item)
        
        folder = self._homerun_folders[name]
        
        # Save expansion state
        was_expanded = folder.isExpanded()
        
        # Clear existing children
        folder.takeChildren()
        
        # Traverse to find connected components (BFS)
        visited = {item}
        queue = [item]
        connected_items = []
        
        while queue:
            current = queue.pop(0)
            if current != item:
                connected_items.append(current)
            
            for neighbor in current.connections:
                if neighbor not in visited:
                    visited.add(neighbor)
                    queue.append(neighbor)
        
        # Add to tree
        for comp in connected_items:
            # Determine icon based on type
            icon = "🔹"
            t = comp.comp_type
            if t == "Lighting": icon = "💡"
            elif t == "Receptacle": icon = "🔌"
            elif t == "Motor": icon = "⚙️"
            elif t == "AC": icon = "❄️"
            elif t == "Panel": icon = "📉"
            
            QTreeWidgetItem(folder, [f"{icon} {comp.name}"])
            
        # Restore expansion state
        if was_expanded:
            folder.setExpanded(True)


    # ==================================================================
    # SELECTION
    # ==================================================================
    def _on_selection_changed(self) -> None:
        try:
            if not self.canvas or not self.canvas.scene:
                return
            items = self.canvas.scene.selectedItems()
        except RuntimeError:
            return

        from ui.canvas import ElectricalComponent
        if items and isinstance(items[0], ElectricalComponent):
            self._current_selected_item = items[0]
            self._name_edit.blockSignals(True)
            self._va_edit.blockSignals(True)
            self._name_edit.setText(items[0].name)
            self._va_edit.setText(str(items[0].va))
            self._name_edit.blockSignals(False)
            self._va_edit.blockSignals(False)
            self._sync_data()
        else:
            self._current_selected_item = None
            self._reset_sidebar()

    def _apply_props(self) -> None:
        if not self._current_selected_item:
            return
        try:
            self._current_selected_item.update_data(
                self._name_edit.text(),
                int(self._va_edit.text()) if self._va_edit.text() else 0,
            )
            self._sync_data()
            self._mark_modified()
        except ValueError:
            pass

    def _reset_sidebar(self) -> None:
        self._summary.setText("Select a component\nto view sizing data…")
        self._lbl_acc.setText("Load Accuracy  : –")
        self._lbl_isc.setText("Short Circuit  : –")
        self._lbl_vd.setText("V-Drop Sync    : –")
        self._name_edit.blockSignals(True)
        self._va_edit.blockSignals(True)
        self._name_edit.clear()
        self._va_edit.clear()
        self._name_edit.blockSignals(False)
        self._va_edit.blockSignals(False)

    # ==================================================================
    # SLD
    # ==================================================================
    def _open_sld(self) -> None:
        from ui.canvas import ElectricalComponent
        items = [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        if not items:
            QMessageBox.warning(self, "SLD", "No components on the canvas.")
            return
        sld = []
        for i in items:
            _, b, w, _ = PECCalculator.calculate_load(i.va)
            sld.append({"name": i.name, "breaker": b, "wire": w})
        self._sld_win = SLDViewer(sld)
        self._sld_win.show()

    # ==================================================================
    # USABILITY
    # ==================================================================
    def run_usability_evaluation(self) -> None:
        from ui.canvas import ElectricalComponent
        comps = [i for i in self.canvas.scene.items() if isinstance(i, ElectricalComponent)]
        if not comps:
            QMessageBox.warning(self, "Usability", "No components to evaluate.")
            return

        outlets = [c for c in comps if any(k in c.name for k in ("Outlet","Duplex","GFCI"))]
        score   = max(0, 100 - len(outlets) * 2)
        status  = "✓ PASS" if score >= 60 else "✗ FAIL – reduce receptacle count"
        self._lbl_usa.setText(f"Usability Score: {score} %")
        QMessageBox.information(self, "Usability Result",
            f"Layout Score    : {score} %\n"
            f"Compliance      : {status}\n"
            f"Receptacles     : {len(outlets)}\n"
            f"Total components: {len(comps)}"
        )

    # ==================================================================
    # SHORT-CIRCUIT
    # ==================================================================
    def _calc_isc(self, va) -> float:
        if va == 0:
            return 0.0
        v   = self.project_data["system_voltage"]
        kva = self.project_data["transformer_kva"]
        z   = self.project_data["transformer_z"]
        return round((kva * 1000) / (v * z) / 1000, 2)

    # ==================================================================
    # EXCEL EXPORT
    # ==================================================================
    def export_to_excel(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Export Schedule", "", _EXCEL_FILTER)
        if not path:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Load Schedule"

            hf   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            hfil = PatternFill("solid", fgColor="1C222D")
            ha   = Alignment(horizontal="center", vertical="center")
            bdr  = Border(
                left=Side(style="thin", color="2D3646"),
                right=Side(style="thin", color="2D3646"),
                top=Side(style="thin", color="2D3646"),
                bottom=Side(style="thin", color="2D3646"),
            )

            headers = ["#", "DESCRIPTION", "PANEL", "VOLTAGE", "LOAD (VA)", "AMPS", "WIRE SIZE"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = hf; cell.fill = hfil; cell.alignment = ha; cell.border = bdr

            widths = [5, 28, 10, 12, 12, 10, 14]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w

            df   = Font(name="Calibri", size=9)
            altf = PatternFill("solid", fgColor="F2F4F7")

            for ri in range(self.table.rowCount()):
                row_data = []
                for ci in range(self.table.columnCount()):
                    cell = self.table.item(ri, ci)
                    row_data.append(cell.text() if cell else "")
                ws.append(row_data)

                er = ri + 2
                for ci in range(1, len(headers) + 1):
                    c = ws.cell(er, ci)
                    c.font = df; c.border = bdr
                    if ri % 2 == 1:
                        c.fill = altf

            # footer
            fr = self.table.rowCount() + 3
            for offset, key in enumerate(["PROJECT:", "ENGINEER:", "STANDARD:"]):
                ws.cell(fr + offset, 1, key).font = Font(bold=True, size=9)
            ws.cell(fr,   2, self.project_data["name"])
            ws.cell(fr+1, 2, self.project_data["author"])
            ws.cell(fr+2, 2, self.project_data["standard"])

            wb.save(path)
            self.statusBar().showMessage(f"📊  Excel saved → {path}")
            QMessageBox.information(self, "Export OK", "Load schedule exported successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed:\n\n{e}")

    # ==================================================================
    # SETTINGS / ABOUT
    # ==================================================================
    def _open_settings(self) -> None:
        dlg = ProjectSettingsDialog(self.project_data, self)
        if dlg.exec():
            self.project_data.update(dlg.get_settings())
            self._sync_data()

    @staticmethod
    def _show_about() -> None:
        QMessageBox.about(
            None, "About ELECDRAFT PRO",
            "<h2 style='color:#00e5ff;'>ELECDRAFT PRO</h2>"
            "<p>Professional Electrical CAD Software</p>"
            "<p><b>Standard:</b> Philippine Electrical Code (PEC)</p>"
            "<p><b>Supported imports:</b> DXF · DWG · PNG · JPG · BMP · TIFF</p>"
            "<p style='color:#888;font-size:11px;'>© 2025 ELECDRAFT – All Rights Reserved</p>"
        )

    # ==================================================================
    # STYLESHEET
    # ==================================================================
    def _apply_stylesheet(self) -> None:
        self.setStyleSheet(f"""
            QMainWindow, QWidget {{
                background-color: {CLR_BG_MAIN};
                color: {CLR_TEXT_BODY};
                font-family: 'Segoe UI', 'Helvetica Neue', sans-serif;
            }}
            #sidePanel {{
                background-color: {CLR_BG_PANEL};
                border-right: 1px solid {CLR_BORDER_LIGHT};
            }}
            #tablePanel {{
                background-color: #111418;
                border-top: 2px solid {CLR_ACCENT};
            }}
            QLabel {{ font-size:10px; color:{CLR_TEXT_DIM}; padding:1px; }}

            #componentTile {{
                background-color: rgba(28,34,45,200);
                border: 1px solid #34445c;
                border-radius: 5px;
            }}
            #componentTile:hover  {{ background-color:#252e3e; border-color:{CLR_ACCENT}; }}
            #componentTile:pressed {{ background-color:#1a2030; }}

            QLineEdit {{
                background-color: #090b0f;
                border: 1px solid {CLR_BORDER};
                color: #fff;
                padding: 7px 10px;
                border-radius: 3px;
                font-size: 11px;
            }}
            QLineEdit:focus {{ border-color:{CLR_ACCENT}; }}

            QTreeWidget, QTableWidget {{
                background-color: {CLR_BG_MAIN};
                border: none;
                color: {CLR_TEXT_BRIGHT};
                font-size: 11px;
            }}
            QTreeWidget::item:selected,
            QTableWidget::item:selected {{
                background-color: rgba(0,229,255,0.12);
                color: #fff;
            }}
            QHeaderView::section {{
                background-color: {CLR_BG_CARD};
                color: {CLR_ACCENT};
                border: 1px solid {CLR_BG_MAIN};
                padding: 6px;
                font-weight: bold;
                font-size: 10px;
            }}

            #exportButton {{
                background-color: {CLR_BG_CARD};
                color: #fff;
                font-weight: bold;
                font-size: 10px;
                padding: 6px 14px;
                border-radius: 4px;
                border: 1px solid {CLR_BORDER};
            }}
            #exportButton:hover {{ border-color:{CLR_ACCENT}; background-color:#252e3e; }}

            #toolButton {{
                background-color: {CLR_BG_CARD};
                color: {CLR_ACCENT};
                font-weight: bold;
                font-size: 10px;
                border: 1px solid {CLR_BORDER};
                padding: 7px;
                border-radius: 4px;
                margin-top: 4px;
            }}
            #toolButton:hover {{ background-color:#252e3e; border-color:{CLR_ACCENT}; }}

            QTabBar::tab {{
                background: {CLR_BG_PANEL};
                padding: 10px 22px;
                border: 1px solid {CLR_BORDER_LIGHT};
                margin-right: 2px;
                border-radius: 4px 4px 0 0;
                color: {CLR_TEXT_DIM};
                font-size: 11px;
            }}
            QTabBar::tab:selected {{ background:{CLR_BG_CARD}; color:{CLR_ACCENT}; border-bottom:2px solid {CLR_ACCENT}; }}
            QTabBar::tab:hover   {{ background:#1e2530; }}

            QGroupBox {{
                border: 1px solid {CLR_BORDER};
                margin-top: 12px;
                padding-top: 8px;
                color: {CLR_ACCENT};
                font-size: 11px;
                border-radius: 4px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 6px;
            }}

            #summaryBox {{
                background-color: #0a0c10;
                color: {CLR_ACCENT};
                padding: 12px 14px;
                border-radius: 4px;
                border: 1px solid {CLR_BORDER};
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 11px;
            }}

            QMenuBar {{
                background-color: #1a1f26;
                color: {CLR_ACCENT};
                font-size: 12px;
                padding: 2px 6px;
            }}
            QMenuBar::item:selected {{ background-color:#252e3e; border-radius:3px; }}
            QMenu {{
                background-color: {CLR_BG_PANEL};
                border: 1px solid {CLR_BORDER};
                border-radius: 4px;
                color: {CLR_TEXT_BRIGHT};
            }}
            QMenu::item:selected {{ background-color:rgba(0,229,255,0.10); }}

            QSplitter::handle       {{ background:{CLR_BG_MAIN}; }}
            QSplitter::handle:hover {{ background:{CLR_BORDER}; }}

            QScrollBar:vertical {{
                background: {CLR_BG_MAIN};
                width: 8px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical {{
                background: {CLR_BORDER};
                border-radius: 4px;
                min-height: 28px;
            }}
            QScrollBar::handle:vertical:hover {{ background:{CLR_ACCENT}; }}
            QScrollBar::add-line, QScrollBar::sub-line {{ height:0; }}

            QStatusBar {{
                background: #12151b;
                border-top: 1px solid {CLR_BORDER_LIGHT};
                color: {CLR_TEXT_BODY};
                font-size: 10px;
                padding: 3px 8px;
            }}
        """)


# ==========================================================================
# ENTRY POINT
# ==========================================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)

    from ui.splash_screen import EnhancedSplash
    splash = EnhancedSplash(LOGO_PATH)
    splash.show()
    app.processEvents()

    window = ElecDraftApp(splash=splash)

    QTimer.singleShot(3000, lambda: splash.finish_loading(window))
    QTimer.singleShot(3500, window.fade_in)

    sys.exit(app.exec())